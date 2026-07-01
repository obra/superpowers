#!/usr/bin/env python3
"""
build_test_case_xlsx.py — Xuất test case đã sinh ra thành file .xlsx đa-sheet
đúng schema 14 cột quan sát được từ file thật TEAM1-697 (canonical), mỗi
sheet tương ứng một màn/sub-feature, sẵn sàng nộp vào quy trình QA hiện có.

Input: 1 file JSON dạng
{
  "sheets": [
    {
      "name": "TEAM2-XXX_Màn danh sách",
      "rows": [
        {
          "tc_id": "STD-01-001",
          "nhom_chuc_nang": "TEAM2-XXX - US-STD-01",
          "module": "Status tiles",
          "muc_tieu": "Load danh sách mặc định",
          "tien_dieu_kien": "Có dữ liệu...",
          "buoc_thuc_hien": "1. ... 2. ...",
          "du_lieu": "...",
          "ket_qua_mong_doi": "...",
          "loai_tc": "Smoke",
          "priority": "High"
        }
      ]
    }
  ]
}

Usage:
    python build_test_case_xlsx.py <input.json> <output.xlsx>

Cột "Kết quả Auto mt DEV", "Thời điểm Auto", "Note Auto", "Kết quả manual mt
DEV" luôn để trống — đó là việc của QA sau khi chạy test, không phải lúc
generate.
"""
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = [
    "TC ID",
    "Nhóm chức năng",
    "Module",
    "Mục tiêu kiểm thử",
    "Tiền điều kiện",
    "Bước thực hiện",
    "Dữ liệu kiểm thử",
    "Kết quả mong đợi",
    "Loại TC",
    "Priority",
    "Kết quả Auto mt DEV",
    "Thời điểm Auto",
    "Note Auto",
    "Kết quả manual mt DEV",
]

ROW_KEYS = [
    "tc_id",
    "nhom_chuc_nang",
    "module",
    "muc_tieu",
    "tien_dieu_kien",
    "buoc_thuc_hien",
    "du_lieu",
    "ket_qua_mong_doi",
    "loai_tc",
    "priority",
]

COL_WIDTHS = [16, 22, 18, 32, 24, 38, 22, 38, 14, 10, 14, 16, 24, 18]

EXCEL_SHEET_NAME_MAX = 31
INVALID_SHEET_CHARS = r'[\\/*?:\[\]]'

# Cột 11-14 (0-based 10..13) = kết quả QA điền TAY sau khi chạy — PHẢI giữ khi regenerate.
TRACKING_IDX = (10, 11, 12, 13)
OBSOLETE_PREFIX = "[KHÔNG CÒN Ở NGUỒN] "


def safe_sheet_name(name: str, used: set) -> str:
    """Excel: tối đa 31 ký tự, không chứa \\/*?:[] , và không trùng tên trong cùng workbook."""
    cleaned = re.sub(INVALID_SHEET_CHARS, "-", name).strip()
    base = cleaned[:EXCEL_SHEET_NAME_MAX]
    candidate = base
    i = 2
    while candidate.lower() in used:
        suffix = f"~{i}"
        candidate = base[: EXCEL_SHEET_NAME_MAX - len(suffix)] + suffix
        i += 1
    used.add(candidate.lower())
    return candidate


def load_existing(path):
    """Đọc .xlsx cũ (nếu có) -> {sheet_name: {tc_id: [14 giá trị]}} để merge."""
    if not path.exists():
        return {}
    wb = load_workbook(path)
    out = {}
    for ws in wb.worksheets:
        rows = {}
        for r in ws.iter_rows(min_row=2, max_col=len(HEADERS), values_only=True):
            tc = r[0]
            if tc is None or str(tc).strip() == "":
                continue
            vals = list(r) + [""] * (len(HEADERS) - len(r))
            rows[str(tc)] = ["" if v is None else v for v in vals]
        out[ws.title] = rows
    return out


def build_workbook(data, existing=None):
    """Regenerate cột 1-10 từ nguồn; GIỮ cột 11-14 (kết quả tester) theo TC ID.
    Row có ở file cũ mà không còn ở nguồn -> giữ lại, đánh dấu OBSOLETE (không xoá)."""
    existing = existing or {}
    wb = Workbook()
    wb.remove(wb.active)  # bỏ sheet mặc định, tự tạo theo data

    used_names = set()
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    wrap = Alignment(wrap_text=True, vertical="top")
    stats = {"preserved": 0, "obsolete": 0}

    for sheet_def in data.get("sheets", []):
        name = safe_sheet_name(sheet_def.get("name", "Sheet"), used_names)
        ws = wb.create_sheet(name)
        old_rows = dict(existing.get(name, {}))  # pop dần khi khớp TC ID

        for col_idx, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
        ws.freeze_panes = "A2"

        row_idx = 2
        for row in sheet_def.get("rows", []):
            for col_idx, key in enumerate(ROW_KEYS, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(key, "")).alignment = wrap
            old = old_rows.pop(str(row.get("tc_id", "")), None)
            if old is not None:  # GIỮ cột 11-14 tester đã điền
                for ci in TRACKING_IDX:
                    ws.cell(row=row_idx, column=ci + 1, value=old[ci]).alignment = wrap
                if any(str(old[ci]).strip() for ci in TRACKING_IDX):
                    stats["preserved"] += 1
            row_idx += 1

        # row cũ không còn ở nguồn (case bị xoá HOẶC tester tự thêm) -> giữ, đánh dấu
        for _tc, vals in old_rows.items():
            for ci in range(len(HEADERS)):
                v = vals[ci]
                if ci == 3 and v and not str(v).startswith(OBSOLETE_PREFIX):
                    v = OBSOLETE_PREFIX + str(v)
                ws.cell(row=row_idx, column=ci + 1, value=v).alignment = wrap
            stats["obsolete"] += 1
            row_idx += 1

        for col_idx, width in enumerate(COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    if not wb.sheetnames:
        wb.create_sheet("Sheet1")

    return wb, stats


def main():
    if len(sys.argv) != 3:
        print("Usage: python build_test_case_xlsx.py <input.json> <output.xlsx>")
        sys.exit(1)

    in_path, out_path = Path(sys.argv[1]), Path(sys.argv[2])
    data = json.loads(in_path.read_text(encoding="utf-8"))
    existing = load_existing(out_path)  # merge nếu file đã tồn tại (giữ cột 11-14)
    wb, stats = build_workbook(data, existing)
    wb.save(out_path)

    total_rows = sum(len(s.get("rows", [])) for s in data.get("sheets", []))
    print(
        json.dumps(
            {
                "output": str(out_path),
                "num_sheets": len(wb.sheetnames),
                "sheet_names": wb.sheetnames,
                "total_test_cases": total_rows,
                "merged": bool(existing),
                "preserved_result_rows": stats["preserved"],
                "obsolete_rows": stats["obsolete"],
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
